import json
import re
import requests
import datetime
from openpyxl import Workbook, load_workbook
import graficas

url = "https://disease.sh/v3/covid-19/historical/all?lastdays=all"
response = requests.get(url)

if response.status_code == 200:
    data = response.json()
    try:
        with open("Data_Base.json", "w") as archivo:
            json.dump(data, archivo)
    except FileNotFoundError:
        print("No se encontro el archivo")
else:
    print("Hubo un error al consultar el API")

try:
    with open("Data_Base.json", "r") as archivo:
        jason = json.load(archivo)
except FileNotFoundError:
    print("No se encontro el archivo")

def crear_libro(decision):
    libro = Workbook()
    hoja = libro.active
    libro.save("Info_Covid.xlsx")
    
def validar_fecha(fecha):
    try:
        mes, dia, año = map(int, fecha.split('/'))
        date_valida = datetime.date(año, mes, dia)
        date_minima = datetime.date(2020, 1, 22)  
        date_maxima = datetime.date(2023, 3, 9)   

        if date_minima <= date_valida <= date_maxima:
            return "{}/{}/{}".format(date_valida.month, date_valida.day, str(date_valida.year)[-2:])
        else:
            print("\nLa fecha está fuera del rango permitido.\n")
    except ValueError:
        print("La fecha ingresada es inválida.")

def calcular_dia(fecha, tipo):
    mes, dia, año = map(int, fecha.split('/'))
    date = datetime.date(año, mes, dia)       
    if mes == 1 and dia == 1:
        mes_anterior = 12
        año_anterior = año - 1
        dia_anterior = 31
        date_anterior = datetime.date(año_anterior, mes_anterior, dia_anterior)
    elif mes == 3 and año == 2020 and dia == 1:
        dia_anterior = 29
        mes_anterior = mes - 1
        date_anterior = datetime.date(año, mes_anterior, dia_anterior)
    elif mes == 3 and dia == 1:
        dia_anterior = 28
        mes_anterior = mes - 1
        date_anterior = datetime.date(año, mes_anterior, dia_anterior)
    elif dia == 1:
        dia_anterior = 30
        mes_anterior = mes - 1
        date_anterior = datetime.date(año, mes_anterior, dia_anterior)
    elif mes == 4 or mes == 6 or mes == 9 or mes == 11 and dia == 1:
        dia_anterior = 31
        mes_anterior = mes - 1
        date_anterior = datetime.date(año, mes_anterior, dia_anterior)
    elif (mes == 3 or mes == 5 or mes == 7 or mes == 8 or mes == 10 or mes == 12) and dia == 1:
        dia_anterior = 30
        mes_anterior = mes - 1
        date_anterior = datetime.date(año, mes_anterior, dia_anterior)
    else:
        dia_anterior = dia - 1
        date_anterior = datetime.date(año, mes, dia_anterior)
    date = "{}/{}/{}".format(date.month, date.day, str(date.year)[-2:])
    date_anterior = "{}/{}/{}".format(date_anterior.month, date_anterior.day, 
                                                  str(date_anterior.year)[-2:])
    casos_dia:int = jason[tipo][date]
    cas_dia:int = jason[tipo][date_anterior]
    total_dia = casos_dia - cas_dia
    return total_dia

def calcular_mes(año, mes, dia, tipo):
    if mes == 2 and año == 2020:
        dia = 29
        dia_anterior = 31
    elif mes == 2:
        dia = 28
        dia_anterior = 31
    elif mes == 3 and año == 2020:
        dia_anterior = 29
    elif mes == 3 and año == 2023:
        dia = 9
        dia_anterior = 28
    elif mes == 3:
        dia_anterior = 28
    elif mes == 4 or mes == 6 or mes == 9 or mes == 11:
        dia = 30
        dia_anterior = 31
    else:
        dia_anterior = 30     
    if mes == 1:
        mes_anterior = 12
        año_anterior = año - 1
        date = datetime.date(año, mes, dia)
        date_anterior = datetime.date(año_anterior, mes_anterior, dia)
        date = "{}/{}/{}".format(date.month, date.day, str(date.year)[-2:])
        date_anterior = "{}/{}/{}".format(date_anterior.month, date_anterior.day, 
                                                  str(date_anterior.year)[-2:])
        casos_mes:int = jason[tipo][date]
        cas_mes:int = jason[tipo][date_anterior]
        total_mes = casos_mes - cas_mes
        return total_mes
    elif mes == 8:
        dia = 31
        dia_anterior = 31
        mes_anterior = mes - 1
    else:
        mes_anterior = mes - 1
    date = datetime.date(año, mes, dia)
    date_anterior = datetime.date(año, mes_anterior, dia_anterior)
    date = "{}/{}/{}".format(date.month, date.day, str(date.year)[-2:])
    date_anterior = "{}/{}/{}".format(date_anterior.month, date_anterior.day, 
                                                  str(date_anterior.year)[-2:])
    casos_mes:int = jason[tipo][date]
    cas_mes:int = jason[tipo][date_anterior]
    total_mes = casos_mes - cas_mes
    return total_mes

def calcular_año(año, mes, dia, tipo):
    if año == 2020:
        date = datetime.date(año, mes, dia)
        date = "{}/{}/{}".format(date.month, date.day, str(date.year)[-2:])
        return jason[tipo][date]
    elif año == 2023:
        dia = 9
        mes = 3
    elif año < 2020 or año > 2023:
        print("Año no valido")
        menu("si")
        return None
    año_anterior = año - 1
    date = datetime.date(año, mes, dia)
    date_anterior = datetime.date(año_anterior, mes, dia)
    date = "{}/{}/{}".format(date.month, date.day, str(date.year)[-2:])
    date_anterior = "{}/{}/{}".format(date_anterior.month, date_anterior.day, 
                                                  str(date_anterior.year)[-2:])
    casos_año:int = jason[tipo][date]
    cas_año:int = jason[tipo][date_anterior]
    total_año = casos_año - cas_año
    return total_año

def agregar_datos_a_lista(decision, datos):
    if decision == "si":
        global lista_datos
    
        for dato in datos:
            lista_datos.append(dato)
    elif decision == "no":
        return None
    else:
        decision = input("No ingresaste un parametro no valido, ingresalo nuevamente: ")
        guardar_exc(decision)

def guardar_exc(datos):
    try:
        libro = Workbook()
    except FileNotFoundError:
        print("No se encontro el archivo")
    hoja = libro.active
    for fila in datos:
            hoja.append(fila)
    libro.save("Info_Covid.xlsx")

def menu(continuar):
    if continuar == "si":
        print("Bienvenido al menu")
        print("1. Casos")
        print("2. Muertes")
        print("3. Recuperados")
        print("4. Crear libreria")
        print("5. Salir\n")
        eleccion = int(input("Escoje un numero:  "))
        
        if eleccion == 1:
            print("Casos")
            print("1. Casos totales")
            print("2. Casos de un dia")
            print("3. Casos del mes")
            print("4. Casos del año")
            print("5. Graficas\n")
            eleccion = int(input("Escoje un numero:  "))
            if eleccion == 1:
                fechas = input("Ingresa una fecha con este formaro MES/DIA/AÑO: ")
                date = validar_fecha(fechas)
                if date is not None: 
                    dia = jason["cases"][date]
                    print("\nSe habian contagiado", dia, "personas\n")
                    datos = [
                    ["Casos totales"],
                    ["Fecha: ", date],
                    ["Casos: ", dia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    menu("si")
                
            elif eleccion == 2:
                fechas = input("Ingresa una fecha con este formaro MES/DIA/AÑO: ")
                date = validar_fecha(fechas)
                if fechas == "1/22/20":
                    dia = jason["cases"][date]
                    print("\nEse dia hubo", dia, "contagios\n")
                    datos = [
                    ["Casos de un dia"],
                    ["Fecha: ", date],
                    ["Casos: ", dia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                elif date is not None: 
                    casosAlDia = calcular_dia(date, "cases")
                    print("\nEse dia hubo", casosAlDia, "contagios\n")
                    datos = [
                    ["Casos de un dia"],
                    ["Fecha: ", date],
                    ["Casos: ", casosAlDia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    menu("si")

            elif eleccion == 3:
                mes = int(input("Agrega un mes: "))
                año = int(input("Agrega un año: "))
                dia = 31
                if mes == 1 and año == 2020:
                    casosAlMes = jason["cases"]["1/31/20"]
                    print("En el mes hubo", casosAlMes, "casos")
                    datos = [
                    ["Casos del mes"],
                    ["Mes: ", mes],
                    ["Año: ", año],
                    ["Casos: ", casosAlMes]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    casosAlMes = calcular_mes(año, mes, dia, "cases")
                    print("\nEn el mes hubo", casosAlMes, "casos\n")
                    datos = [
                    ["Casos del mes"],
                    ["Mes: ", mes],
                    ["Año: ", año],
                    ["Casos: ", casosAlMes]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")

            elif eleccion == 4:
                año = int(input("Ingrese el año: "))
                mes = 12
                dia = 31
                casosAlAño = calcular_año(año, mes, dia, "cases")
                if casosAlAño is not None:
                    print("\nEn el año hubo", casosAlAño, "casos\n")
                    datos = [
                        ["Casos del año"],
                        ["Año: ", año],
                        ["Casos: ", casosAlAño]
                   ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    exit

            elif eleccion == 5:
                print("Graficas")
                print("1. Casos del mes")
                print("2. Casos del año")
                eleccion = int(input("Escoje un numero:  "))
                if eleccion == 1:
                    año = int(input("Ingrese el año: "))
                    lista_meses = []
                    if 2020 < año < 2023:
                        for mes in range(1,13):
                            dia = 31
                            lista_meses.append(calcular_mes(año, mes, dia, "cases"))
                    elif año == 2023:
                        lista_meses.append(calcular_mes(2023, 1, 31, "cases"))
                        lista_meses.append(calcular_mes(2023, 2, 31, "cases"))
                        lista_meses.append(calcular_mes(2023, 3, 9, "cases"))
                    elif año == 2020:
                         lista_meses.append(jason["cases"]["1/31/20"])
                         for mes in range(2,13):
                            dia = 31
                            lista_meses.append(calcular_mes(año, mes, dia, "cases"))
                    else:
                        print("No ingresaste un parametro no valido, ingresalo nuevamente: ")
                        menu("si")
                    graficas.grafica_mes(lista_meses, "Casos")
                    menu("si")

                elif eleccion == 2:
                    años_validos = [2020, 2021, 2022, 2023]
                    lista_años = []
                    for año in años_validos:
                        dia = 31
                        mes = 12
                        lista_años.append(calcular_año(año,mes,dia,"cases"))
                    graficas.grafica_año(lista_años, "casos")
                    menu("si")

        elif eleccion == 2:
            print("Muertes")
            print("1. Muertes totales")
            print("2. Muertes de un dia")
            print("3. Muertes del mes")
            print("4. Muertes del año")
            print("5. Graficas\n")
            eleccion = int(input("Escoje un numero:  "))
            if eleccion == 1:
                fechas = input("Ingresa una fecha con este formaro MES/DIA/AÑO: ")
                date = validar_fecha(fechas)
                if date is not None: 
                    dia = jason["deaths"][date]
                    print("\nSe habian muerto", dia, "personas\n")
                    datos = [
                    ["Muertos totales"],
                    ["Fecha: ", date],
                    ["Muertos: ", dia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    menu("si")
                
            elif eleccion == 2:
                fechas = input("Ingresa una fecha con este formaro MES/DIA/AÑO: ")
                date = validar_fecha(fechas)
                if fechas == "1/22/20":
                    dia = jason["deaths"][date]
                    print("\nEse dia hubo", dia, "muertos\n")
                    datos = [
                    ["Muertos de un dia"],
                    ["Fecha: ", date],
                    ["Muertos: ", dia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                elif date is not None: 
                    casosAlDia = calcular_dia(date, "deaths")
                    print("\nEse dia hubo", casosAlDia, "muertos\n")
                    datos = [
                    ["Muertos de un dia"],
                    ["Fecha: ", date],
                    ["Muertos: ", casosAlDia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    menu("si")

            elif eleccion == 3:
                mes = int(input("Agrega un mes: "))
                año = int(input("Agrega un año: "))
                dia = 31
                if mes == 1 and año == 2020:
                    casosAlMes = jason["deaths"]["1/31/20"]
                    print("En el mes hubo", casosAlMes, "muetos")
                    datos = [
                    ["Muertos del mes"],
                    ["Mes: ", mes],
                    ["Año: ", año],
                    ["Muertos: ", casosAlMes]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    casosAlMes = calcular_mes(año, mes, dia, "deaths")
                    print("\nEn el mes hubo", casosAlMes, "muertos\n")
                    datos = [
                    ["Muertos del mes"],
                    ["Mes: ", mes],
                    ["Año: ", año],
                    ["Muertos: ", casosAlMes]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")

            elif eleccion == 4:
                año = int(input("Ingrese el año: "))
                mes = 12
                dia = 31
                casosAlAño = calcular_año(año, mes, dia, "deaths")
                print("\nEn el año hubo", casosAlAño, "muertos\n")
                datos = [
                    ["Muertos del año"],
                    ["Año: ", año],
                    ["Muertos: ", casosAlAño]
                ]
                seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                agregar_datos_a_lista(seleccion, datos)
                menu("si")

            elif eleccion == 5:
                print("Graficas")
                print("1. Muertos del mes")
                print("2. Muertos del año")
                eleccion = int(input("Escoje un numero:  "))
                if eleccion == 1:
                    año = int(input("Ingrese el año: "))
                    lista_meses = []
                    if 2020 < año < 2023:
                        for mes in range(1,13):
                            dia = 31
                            lista_meses.append(calcular_mes(año, mes, dia, "deaths"))
                    elif año == 2023:
                        lista_meses.append(calcular_mes(2023, 1, 31, "deaths"))
                        lista_meses.append(calcular_mes(2023, 2, 31, "deaths"))
                        lista_meses.append(calcular_mes(2023, 3, 9, "deaths"))
                    elif año == 2020:
                         lista_meses.append(jason["deaths"]["1/31/20"])
                         for mes in range(2,13):
                            dia = 31
                            print(mes)
                            lista_meses.append(calcular_mes(año, mes, dia, "deaths"))
                    else:
                        print("No ingresaste un parametro no valido, ingresalo nuevamente: ")
                        menu("si")
                    graficas.grafica_mes(lista_meses, "Muertes")
                    menu("si")

                elif eleccion == 2:
                    años_validos = [2020, 2021, 2022, 2023]
                    lista_años = []
                    for año in años_validos:
                        dia = 31
                        mes = 12
                        lista_años.append(calcular_año(año,mes,dia,"deaths"))
                    graficas.grafica_año(lista_años, "Muertos")
                    menu("si")

        elif eleccion == 3:
            print("Recuperados")
            print("1. Recuperados totales")
            print("2. Recuperados de un dia")
            print("3. Recuperados del mes")
            print("4. Recuperados del año")
            print("5. Graficas\n")
            eleccion = int(input("Escoje un numero:  "))
            if eleccion == 1:
                fechas = input("Ingresa una fecha con este formaro MES/DIA/AÑO: ")
                date = validar_fecha(fechas)
                if date is not None: 
                    dia = jason["recovered"][date]
                    print("\nSe habian recuperado", dia, "personas\n")
                    datos = [
                    ["Recuperados totales"],
                    ["Fecha: ", date],
                    ["Recuperados: ", dia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    menu("si")
                
            elif eleccion == 2:
                fechas = input("Ingresa una fecha con este formaro MES/DIA/AÑO: ")
                date = validar_fecha(fechas)
                if fechas == "1/22/20":
                    dia = jason["recovered"][date]
                    print("\nEse dia hubo", dia, "recuperados\n")
                    datos = [
                    ["Recuperados de un dia"],
                    ["Fecha: ", date],
                    ["Recuperados: ", dia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                elif date is not None: 
                    casosAlDia = calcular_dia(date, "recovered")
                    print("\nEse dia hubo", casosAlDia, "recuperados\n")
                    datos = [
                    ["Recuperados de un dia"],
                    ["Fecha: ", date],
                    ["Recuperados: ", casosAlDia]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    menu("si")

            elif eleccion == 3:
                mes = int(input("Agrega un mes: "))
                año = int(input("Agrega un año: "))
                dia = 31
                if mes == 1 and año == 2020:
                    casosAlMes = jason["recovered"]["1/31/20"]
                    print("En el mes hubo", casosAlMes, "recuperados")
                    datos = [
                    ["Recuperados del mes"],
                    ["Mes: ", mes],
                    ["Año: ", año],
                    ["Recuperados: ", casosAlMes]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")
                else:
                    casosAlMes = calcular_mes(año, mes, dia, "recovered")
                    print("\nEn el mes hubo", casosAlMes, "recuperados\n")
                    datos = [
                    ["Recuperados del mes"],
                    ["Mes: ", mes],
                    ["Año: ", año],
                    ["Recuperados: ", casosAlMes]
                ]
                    seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                    agregar_datos_a_lista(seleccion, datos)
                    menu("si")

            elif eleccion == 4:
                año = int(input("Ingrese el año: "))
                mes = 12
                dia = 31
                casosAlAño = calcular_año(año, mes, dia, "recovered")
                print("\nEn el mes hubo", casosAlAño, "recuperados\n")
                datos = [
                    ["Recuperados del año"],
                    ["Año: ", año],
                    ["Recuperados: ", casosAlAño]
                ]
                seleccion = input("Deseas guardar el dato en el libro (si/no): ")
                agregar_datos_a_lista(seleccion, datos)
                menu("si")

            elif eleccion == 5:
                print("Graficas")
                print("1. Recuperados del mes")
                print("2. Recuperados del año")
                eleccion = int(input("Escoje un numero:  "))
                if eleccion == 1:
                    año = int(input("Ingrese el año: "))
                    lista_meses = []
                    if 2020 < año < 2023:
                        for mes in range(1,13):
                            dia = 31
                            lista_meses.append(calcular_mes(año, mes, dia, "recovered"))
                    elif año == 2023:
                        lista_meses.append(calcular_mes(2023, 1, 31, "recovered"))
                        lista_meses.append(calcular_mes(2023, 2, 31, "recovered"))
                        lista_meses.append(calcular_mes(2023, 3, 9, "recovered"))
                    elif año == 2020:
                         lista_meses.append(jason["recovered"]["1/31/20"])
                         for mes in range(2,13):
                            dia = 31
                            lista_meses.append(calcular_mes(año, mes, dia, "recovered"))
                    else:
                        print("No ingresaste un parametro no valido, ingresalo nuevamente: ")
                        menu("si")
                    graficas.grafica_mes(lista_meses, "Recuperados")
                    menu("si")

                elif eleccion == 2:
                    años_validos = [2020, 2021, 2022, 2023]
                    lista_años = []
                    for año in años_validos:
                        dia = 31
                        mes = 12
                        lista_años.append(calcular_año(año,mes,dia,"recovered"))
                    graficas.grafica_año(lista_años, "Recuperados")
                    menu("si")

        elif eleccion == 4:
            crear_libro("si")
            menu("si")
        
        elif eleccion == 5:
            print("Adios, que tengas un buen dia \U0001F60A")
            if not lista_datos:
                exit
            else:
                guardar_exc(lista_datos)
            exit
        else:
            eleccion = input("No ingresaste un parametro no valido, ingresalo nuevamente: ")
            menu(eleccion)
    elif continuar == "no":
        print("Adios, que tengas un buen dia \U0001F60A")
    else:
        continuar = input("No ingresaste un parametro no valido, ingresalo nuevamente: ")
        menu(continuar)

lista_datos = []
